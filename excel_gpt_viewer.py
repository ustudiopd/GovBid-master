import sys, os, json
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, QFrame, QTextBrowser, QTextEdit,
    QSplitter, QLabel, QMessageBox, QStyledItemDelegate
)
from PyQt5.QtGui import QColor, QFont, QPen
from PyQt5.QtCore import Qt
from dotenv import load_dotenv
import openpyxl
import requests

# .env에서 GPT 키/모델 불러오기
load_dotenv()
GPT_API_KEY = os.getenv("CHATGPT_API_KEY", "")
GPT_MODEL = os.getenv("CHATGPT_MODEL", "gpt-4.1-mini")

def apply_tint(hex_rgb, tint):
    # hex_rgb: 'RRGGBB', tint: -1.0~1.0
    ch = [int(hex_rgb[0:2], 16), int(hex_rgb[2:4], 16), int(hex_rgb[4:6], 16)]
    out = [0, 0, 0]
    for i in range(3):
        c = ch[i]
        if tint < 0:
            nc = c * (1 + tint)
        else:
            nc = c * (1 - tint) + 255 * tint
        out[i] = max(0, min(int(round(nc)), 255))
    return QColor(out[0], out[1], out[2])

class BorderDelegate(QStyledItemDelegate):
    """openpyxl border 정보에 따라 셀 테두리만 그려주는 Delegate"""
    def paint(self, painter, option, index):
        # 1) 기본 렌더링 (배경·글자 등)
        super().paint(painter, option, index)
        # 2) border_info에 따라 테두리만 그림
        border_info = index.data(Qt.UserRole)
        if not isinstance(border_info, dict):
            return
        pen = QPen(QColor('#555555'))
        pen.setWidth(1)
        painter.save()
        painter.setPen(pen)
        r = option.rect
        # 위쪽 테두리
        if border_info.get("top"):
            painter.drawLine(r.topLeft(), r.topRight())
        # 아래쪽
        if border_info.get("bottom"):
            painter.drawLine(r.bottomLeft(), r.bottomRight())
        # 왼쪽
        if border_info.get("left"):
            painter.drawLine(r.topLeft(), r.bottomLeft())
        # 오른쪽
        if border_info.get("right"):
            painter.drawLine(r.topRight(), r.bottomRight())
        painter.restore()

class ExcelGPTViewer(QMainWindow):
    def log(self, msg):
        if hasattr(self, 'log_output'):
            self.log_output.append(msg)
        print(msg)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel + GPT 분석기")
        self.setGeometry(100, 100, 1400, 900)
        self.json_path = None
        self.excel_path = None

        # 메인 레이아웃
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # QSplitter(좌:엑셀, 우:챗)
        self.splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(self.splitter)

        # 1) 엑셀 뷰어 패널
        excel_panel = QWidget()
        excel_layout = QVBoxLayout(excel_panel)
        excel_layout.setContentsMargins(8, 8, 8, 8)
        file_btn = QPushButton("엑셀 파일 열기")
        file_btn.clicked.connect(self.open_excel)
        excel_layout.addWidget(file_btn)
        self.excel_view = QTableWidget()
        # 기본 그리드(격자) 끄기
        self.excel_view.setShowGrid(False)
        # 셀별 테두리 그리도록 Delegate 설정
        self.excel_view.setItemDelegate(BorderDelegate(self.excel_view))
        self.excel_view.cellChanged.connect(self.on_cell_changed)
        excel_layout.addWidget(self.excel_view)
        self.splitter.addWidget(excel_panel)

        # 2) 챗봇 패널
        chat_frame = QFrame()
        chat_layout = QVBoxLayout(chat_frame)
        chat_layout.setContentsMargins(8, 8, 8, 8)
        chat_layout.setSpacing(8)
        self.model_label = QLabel(f"모델: {GPT_MODEL}")
        self.model_label.setStyleSheet("font-weight:bold;color:#0057b8;")
        chat_layout.addWidget(self.model_label)
        self.chat_output = QTextBrowser()
        self.chat_output.setOpenExternalLinks(True)
        self.chat_output.setMinimumHeight(180)
        chat_layout.addWidget(self.chat_output)
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setFrameShadow(QFrame.Sunken)
        chat_layout.addWidget(line)
        self.chat_input = QTextEdit()
        self.chat_input.setPlaceholderText("질문을 입력하세요...")
        self.chat_input.setMinimumHeight(40)
        self.chat_input.setStyleSheet("border:2px solid #000;")
        chat_layout.addWidget(self.chat_input)
        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        self.send_btn = QPushButton("질문하기")
        self.send_btn.clicked.connect(self.ask_gpt)
        btn_layout.addWidget(self.send_btn)
        btn_layout.addStretch(1)
        btn_widget = QWidget(); btn_widget.setLayout(btn_layout)
        chat_layout.addWidget(btn_widget)
        self.splitter.addWidget(chat_frame)
        self.splitter.setSizes([900, 500])

        # 로그 메시지창 (하단)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setMaximumHeight(80)
        self.log_output.setStyleSheet("background:#222;color:#eee;font-size:12px;")
        # 메인 레이아웃에 로그창 추가 (세로로 쌓기)
        vbox = QVBoxLayout()
        vbox.setContentsMargins(0,0,0,0)
        vbox.setSpacing(0)
        vbox.addLayout(main_layout)
        vbox.addWidget(self.log_output)
        main_widget.setLayout(vbox)

        # 헤더/그리드 스타일
        self.excel_view.setStyleSheet("""
        QHeaderView::section {
            background-color: #3F4A73;
            color: white;
            font-weight: bold;
            border: 1px solid #555;
        }
        QTableWidget {
            gridline-color: #AAAAAA;
        }
        """)

    def open_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return
        self.log(f"[작업] 엑셀 파일 열기: {path}")
        try:
            self.excel_path = path
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            # accent1 파랑 계열 RGB를 하드코딩
            accent_colors = {1: '3B4E87'}
            rows, cols = ws.max_row, ws.max_column
            table = self.excel_view
            table.blockSignals(True)
            table.clear()
            table.setRowCount(rows)
            table.setColumnCount(cols)

            # 셀 값, 폰트, 배경, 정렬, 테두리(UserRole에 정보 저장)
            from openpyxl.styles.borders import Border
            for r in range(1, rows+1):
                for c in range(1, cols+1):
                    cell = ws.cell(r, c)
                    text = cell.value or ""
                    item = QTableWidgetItem(str(text))

                    # 폰트
                    f = cell.font
                    point_size = int(f.sz) if f.sz is not None else -1
                    qf = QFont(f.name, point_size)
                    qf.setBold(f.b)
                    qf.setItalic(f.i)
                    item.setFont(qf)

                    # 배경색 (Excel 테마 컬러 + 틴트 지원)
                    color = None
                    fill = cell.fill
                    fg = fill.fgColor if hasattr(fill, 'fgColor') else None
                    try:
                        if fill.patternType in ('solid', 'gray125', 'darkGrid', 'lightGrid') and fg:
                            # 1) Theme 컬러(무조건 파랑 accent1) + 틴트
                            if fg.type == 'theme':
                                tint = getattr(fg, 'tint', 0.0)
                                hex_rgb = accent_colors[1]
                                color = apply_tint(hex_rgb, tint)
                                self.log(f"[THEME+TINT:파랑] ({r},{c}) tint={tint} {hex_rgb} → ({color.red()},{color.green()},{color.blue()}) 적용")
                            # 2) RGB 컬러
                            elif fg.type == 'rgb' and fg.rgb:
                                rgb = fg.rgb[2:] if fg.rgb.startswith('FF') else fg.rgb  # e.g. 'FFFFCC00' or 'FFCC00'
                                color = QColor(int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16))
                                self.log(f"[RGB] ({r},{c}) {rgb} → ({color.red()},{color.green()},{color.blue()}) 적용")
                            # 3) Indexed 컬러
                            elif fg.type == 'indexed' and fg.indexed is not None:
                                from openpyxl.styles.colors import COLOR_INDEX
                                idx = fg.indexed
                                if 0 <= idx < len(COLOR_INDEX):
                                    hexcol = COLOR_INDEX[idx][2:]  # 'RRGGBB'
                                    color = QColor(int(hexcol[0:2],16), int(hexcol[2:4],16), int(hexcol[4:6],16))
                                    self.log(f"[INDEXED] ({r},{c}) idx={idx} {hexcol} → ({color.red()},{color.green()},{color.blue()}) 적용")
                        # 4) Gradient Fill (첫 stop만 사용)
                        elif hasattr(fill, 'gradientType') and fill.gradientType:
                            stops = getattr(fill, 'stop', None)
                            if stops and hasattr(stops[0], 'color') and hasattr(stops[0].color, 'rgb'):
                                rgb = stops[0].color.rgb[2:] if stops[0].color.rgb.startswith('FF') else stops[0].color.rgb
                                color = QColor(int(rgb[0:2],16), int(rgb[2:4],16), int(rgb[4:6],16))
                                self.log(f"[GRADIENT] ({r},{c}) {rgb} → ({color.red()},{color.green()},{color.blue()}) 적용")
                    except Exception as e:
                        self.log(f"[ERROR] ({r},{c}) 색상 파싱 오류: {e}")
                    if color and (color.red(), color.green(), color.blue()) != (255, 255, 255):
                        item.setBackground(color)

                    # 정렬
                    align = cell.alignment
                    qt_align = 0
                    if align.horizontal == 'center': qt_align |= Qt.AlignHCenter
                    elif align.horizontal == 'right': qt_align |= Qt.AlignRight
                    else: qt_align |= Qt.AlignLeft
                    if align.vertical == 'center': qt_align |= Qt.AlignVCenter
                    elif align.vertical == 'bottom': qt_align |= Qt.AlignBottom
                    else: qt_align |= Qt.AlignTop
                    item.setTextAlignment(qt_align)

                    # openpyxl Border 객체를 dict로 변환해 UserRole에 저장
                    b = cell.border
                    border_info = {
                        "top":    bool(b.top and b.top.style),
                        "bottom": bool(b.bottom and b.bottom.style),
                        "left":   bool(b.left and b.left.style),
                        "right":  bool(b.right and b.right.style),
                    }
                    item.setData(Qt.UserRole, border_info)
                    table.setItem(r-1, c-1, item)

            # 병합 셀
            for merged in ws.merged_cells.ranges:
                r0, c0 = merged.min_row-1, merged.min_col-1
                rs = merged.max_row - merged.min_row + 1
                cs = merged.max_col - merged.min_col + 1
                table.setSpan(r0, c0, rs, cs)

            # 열 너비
            for idx, col_dim in ws.column_dimensions.items():
                col = openpyxl.utils.column_index_from_string(idx) - 1
                if col < cols and col_dim.width:
                    table.setColumnWidth(col, int(col_dim.width * 7))
            # 행 높이
            for r, row_dim in ws.row_dimensions.items():
                if row_dim.height is not None and r-1 < rows:
                    table.setRowHeight(r-1, int(row_dim.height * 1.2))

            # 금액 열 읽기전용(예시: 5번째 열)
            amount_col = 4
            for r in range(rows):
                it = table.item(r, amount_col)
                if it:
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)

            table.blockSignals(False)

            # JSON 파일 경로
            json_path = os.path.splitext(path)[0] + ".json"
            self.json_path = json_path
            # 초기 JSON 저장
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(self._widget_to_json_schema(), f, ensure_ascii=False, indent=2)
            self.log(f"[작업] JSON 파일 저장: {json_path}")
        except Exception as e:
            self.log(f"[오류] 엑셀 파일 열기 실패: {e}")
            QMessageBox.critical(self, "엑셀 파일 오류", f"엑셀 파일을 불러올 수 없습니다:\n{e}")

    def on_cell_changed(self, row, col):
        if not self.json_path:
            return
        data_dict = self._widget_to_json_schema()
        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=2)
        self.log(f"[작업] 셀 변경: ({row},{col}) → JSON 동기화")

    def _widget_to_json_schema(self):
        table = self.excel_view
        result = {
            "meta": {},
            "items": [],
            "discounts": [],
            "summary": {},
            "comments": ""
        }
        current_category = None
        for row in range(table.rowCount()):
            a_item = table.item(row, 0)
            d_item = table.item(row, 3)
            a = a_item.text() if a_item else ""
            d = d_item.text() if d_item else ""
            # 1) 섹션 헤더
            if a and not d:
                current_category = a.strip()
                continue
            # 2) 품목 행
            if a and d:
                try:
                    item = {
                        "category": current_category,
                        "description": a.strip(),
                        "unit_price": float(table.item(row, 1).text()) if table.item(row, 1) and table.item(row, 1).text() else 0,
                        "quantity": float(table.item(row, 2).text()) if table.item(row, 2) and table.item(row, 2).text() else 0,
                        "unit_count": float(table.item(row, 3).text()),
                        "amount": float(table.item(row, 4).text()) if table.columnCount() > 4 and table.item(row, 4) and table.item(row, 4).text() else None
                    }
                    result["items"].append(item)
                except Exception:
                    continue
                continue
            # 3) 요약/할인/합계 행 (A열 비어있고 D열에 숫자)
            if not a and d:
                try:
                    val = float(d)
                    if val < 0:
                        result["discounts"].append({"description": "", "amount": -val})
                    else:
                        if "subtotal" not in result["summary"]:
                            result["summary"]["subtotal"] = val
                        elif "tax_amount" not in result["summary"]:
                            result["summary"]["tax_amount"] = val
                        else:
                            result["summary"]["total_due"] = val
                except Exception:
                    continue
                continue
        return result

    def ask_gpt(self):
        user_q = self.chat_input.toPlainText().strip()
        if not user_q or not self.json_path:
            return
        with open(self.json_path, "r", encoding="utf-8") as f:
            quotation_json = f.read()
        messages = [
            {"role": "system", "content": "아래 견적서 JSON을 참고해 질문에 답변해 주세요."},
            {"role": "user", "content": f"견적서 데이터:\n```json\n{quotation_json}\n```\n질문: {user_q}"}
        ]
        answer = ask_gpt_api(messages, GPT_API_KEY, GPT_MODEL)
        self.chat_output.append(f"<b>질문:</b> {user_q}")
        self.chat_output.append(f"<b>GPT:</b>")
        self.chat_output.append(answer)
        self.chat_input.clear()
        self.log(f"[작업] GPT 질문 전송 및 응답 수신 완료")

def ask_gpt_api(messages, api_key, model):
    if not api_key:
        return "[OpenAI API 키를 .env에 입력하세요]"
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": model,
        "messages": messages,
        "max_tokens": 2048,
        "temperature": 0.7
    }
    try:
        resp = requests.post(url, headers=headers, json=data, timeout=30)
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"[GPT 호출 오류] {e}"

if __name__ == "__main__":
    app = QApplication(sys.argv)
    viewer = ExcelGPTViewer()
    viewer.show()
    sys.exit(app.exec_()) 